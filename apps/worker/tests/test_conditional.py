from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from sheet_diff.conditional import extract_conditional_formatting
from sheet_diff.engine import compare_workbooks
from sheet_diff.models import DiffOptions

SAMPLES = Path(__file__).resolve().parents[3] / "samples"


def test_conditional_uses_sheet_name_not_file_stem(tmp_path):
    before = tmp_path / "cf_before.xlsx"
    after = tmp_path / "cf_after.xlsx"

    for path, rules in [(before, []), (after, [CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill())])]:
        wb = Workbook()
        ws = wb.active
        ws.title = "MySheet"
        ws["A1"] = 1
        if rules:
            ws.conditional_formatting.add("A1:A1", rules[0])
        wb.save(path)

    b_cf = extract_conditional_formatting(before)
    a_cf = extract_conditional_formatting(after)
    assert "MySheet" in b_cf or "MySheet" in a_cf
    assert "sheet1" not in b_cf and "sheet1" not in a_cf

    result = compare_workbooks(before, after, DiffOptions())
    cf_changes = [c for c in result.cells if c.kind == "style" and c.cell == ""]
    if cf_changes:
        assert cf_changes[0].sheet == "MySheet"
