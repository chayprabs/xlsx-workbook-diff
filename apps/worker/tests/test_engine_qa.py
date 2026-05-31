"""Round 2 engine QA — explicit coverage for compare_workbooks edge cases."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from sheet_diff.engine import compare_workbooks
from sheet_diff.models import DiffOptions
from sheet_diff.reports import KIND_COLORS, write_diff_workbook

SAMPLES = Path(__file__).resolve().parents[3] / "samples"
INVOICE_BEFORE = SAMPLES / "invoice_before.xlsx"
INVOICE_AFTER = SAMPLES / "invoice_after.xlsx"


def test_identical_workbooks_zero_changes():
    before = INVOICE_BEFORE
    result = compare_workbooks(before, before, DiffOptions())
    assert result.cells == []
    assert result.summary.total_changes == 0


def test_pricing_structure_summary_in_added():
    result = compare_workbooks(
        SAMPLES / "pricing_before.xlsx",
        SAMPLES / "pricing_after.xlsx",
        DiffOptions(),
    )
    assert "Summary" in result.structure.sheets.get("added", [])


def test_financial_named_range_revenue_range_changed():
    result = compare_workbooks(
        SAMPLES / "financial_before.xlsx",
        SAMPLES / "financial_after.xlsx",
        DiffOptions(),
    )
    changed = {c["name"]: c for c in result.named_ranges.changed}
    assert "RevenueRange" in changed
    assert changed["RevenueRange"]["before"] == "Report!$B$2:$B$4"
    assert changed["RevenueRange"]["after"] == "Report!$B$2:$B$5"


def test_financial_chart_diff_non_empty_on_report_sheet():
    result = compare_workbooks(
        SAMPLES / "financial_before.xlsx",
        SAMPLES / "financial_after.xlsx",
        DiffOptions(),
    )
    assert len(result.charts) >= 1
    assert result.charts[0].sheet == "Report"
    assert result.charts[0].diff["before"]["title"] == "Revenue"
    assert result.charts[0].diff["after"]["title"] == "Revenue Q1"


def test_style_only_diff(tmp_path):
    before = tmp_path / "style_before.xlsx"
    after = tmp_path / "style_after.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["A1"].font = Font(bold=False)
    wb.save(before)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2["A1"] = "Hello"
    ws2["A1"].font = Font(bold=True)
    wb2.save(after)

    result = compare_workbooks(before, after, DiffOptions())
    assert len(result.cells) == 1
    assert result.cells[0].kind == "style"
    assert result.cells[0].cell == "A1"


def test_invoice_added_row_a4_kind_is_added():
    before, after = INVOICE_BEFORE, INVOICE_AFTER
    result = compare_workbooks(before, after, DiffOptions())
    by_cell = {c.cell: c for c in result.cells if c.sheet == "Invoice"}
    assert by_cell["A4"].kind == "added"
    assert by_cell["B4"].kind == "added"
    assert "value" not in {by_cell["A4"].kind, by_cell["B4"].kind}


def test_bump_counts_formula_style_composite(tmp_path):
    before = tmp_path / "fs_before.xlsx"
    after = tmp_path / "fs_after.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = "=1+2"
    ws["A1"].font = Font(bold=False)
    wb.save(before)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Calc"
    ws2["A1"] = "=1+3"
    ws2["A1"].font = Font(bold=True)
    wb2.save(after)

    result = compare_workbooks(before, after, DiffOptions())
    assert len(result.cells) == 1
    assert result.cells[0].kind == "formula+style"
    counts = result.summary.by_sheet["Calc"]
    assert counts["formula"] == 1
    assert counts["style"] == 1
    assert counts["total"] == 1
    assert counts["value"] == 0


def test_write_diff_workbook_composite_kind_colors(tmp_path):
    before, after = INVOICE_BEFORE, INVOICE_AFTER
    result = compare_workbooks(before, after, DiffOptions())
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(result, before, after, out)

    wb = load_workbook(out)
    ws = wb["Invoice"]
    def expected_rgb(kind: str) -> str:
        color = KIND_COLORS.get(kind)
        if color:
            return "00" + color.replace("#", "")
        for key in ("formula", "style", "value", "added", "removed"):
            if key in kind:
                return "00" + KIND_COLORS[key].replace("#", "")
        return "00" + KIND_COLORS["value"].replace("#", "")

    for change in result.cells:
        fill_rgb = ws[change.cell].fill.start_color.rgb
        assert fill_rgb == expected_rgb(change.kind), (
            f"{change.cell} kind={change.kind}: got {fill_rgb}, expected {expected_rgb(change.kind)}"
        )

    # formula+style composite from synthetic pair
    fs_before = tmp_path / "fs_before.xlsx"
    fs_after = tmp_path / "fs_after.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "=1+2"
    ws["A1"].font = Font(bold=False)
    wb.save(fs_before)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "S"
    ws2["A1"] = "=1+3"
    ws2["A1"].font = Font(bold=True)
    wb2.save(fs_after)
    fs_result = compare_workbooks(fs_before, fs_after, DiffOptions())
    fs_out = tmp_path / "fs_diff.xlsx"
    write_diff_workbook(fs_result, fs_before, fs_after, fs_out)
    fs_fill = load_workbook(fs_out)["S"]["A1"].fill.start_color.rgb
    assert fs_fill == "00" + KIND_COLORS["formula+style"].replace("#", "")
