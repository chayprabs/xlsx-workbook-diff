from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from sheet_diff.charts import diff_charts, extract_charts_from_xlsx
from sheet_diff.conditional import diff_conditional, extract_conditional_formatting
from sheet_diff.formula import classify_formula_change, formula_ast_diff
from sheet_diff.models import (
    CellChange,
    ChartDiff,
    DiffOptions,
    DiffResult,
    DiffSummary,
    HiddenDiff,
    NamedRangesDiff,
    StructureDiff,
)
from sheet_diff.structure import (
    diff_hidden,
    diff_named_ranges,
    diff_sheet_structure,
    diff_tables,
    sheet_structure,
)
from sheet_diff.style import style_diff, style_snapshot
from sheet_diff.tolerance import resolve_tolerance, values_equal


def _cell_coord(cell: Cell) -> str:
    return f"{get_column_letter(cell.column)}{cell.row}"


def _iter_cells(ws, max_row: int | None = None, max_col: int | None = None):
    mr = max_row or ws.max_row or 1
    mc = max_col or ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            yield cell


def compare_workbooks(
    before_path: Path,
    after_path: Path,
    opts: DiffOptions | None = None,
) -> DiffResult:
    opts = opts or DiffOptions()
    abs_tol, rel_tol = resolve_tolerance(opts)

    wb_b = load_workbook(before_path, data_only=False, read_only=False)
    wb_a = load_workbook(after_path, data_only=False, read_only=False)
    wb_b_vals = load_workbook(before_path, data_only=True, read_only=False)
    wb_a_vals = load_workbook(after_path, data_only=True, read_only=False)

    cells: list[CellChange] = []
    by_sheet: dict[str, dict[str, int]] = {}

    def bump(sheet: str, kind: str):
        by_sheet.setdefault(sheet, {"value": 0, "formula": 0, "style": 0, "total": 0})
        by_sheet[sheet]["total"] += 1
        parts = kind.split("+")
        if any(p in parts for p in ("value", "added", "removed")):
            by_sheet[sheet]["value"] += 1
        if "formula" in parts:
            by_sheet[sheet]["formula"] += 1
        if "style" in parts:
            by_sheet[sheet]["style"] += 1

    all_sheets = sorted(set(wb_b.sheetnames) | set(wb_a.sheetnames))
    for sheet in all_sheets:
        if sheet not in wb_b.sheetnames:
            ws_a = wb_a[sheet]
            for cell in _iter_cells(ws_a):
                if cell.value is None and cell.data_type != "f":
                    continue
                cells.append(
                    CellChange(
                        sheet=sheet,
                        cell=_cell_coord(cell),
                        kind="added",
                        after=cell.value,
                    )
                )
                bump(sheet, "added")
            continue
        if sheet not in wb_a.sheetnames:
            ws_b = wb_b[sheet]
            for cell in _iter_cells(ws_b):
                if cell.value is None and cell.data_type != "f":
                    continue
                cells.append(
                    CellChange(
                        sheet=sheet,
                        cell=_cell_coord(cell),
                        kind="removed",
                        before=cell.value,
                    )
                )
                bump(sheet, "removed")
            continue

        ws_b, ws_a = wb_b[sheet], wb_a[sheet]
        ws_bv, ws_av = wb_b_vals[sheet], wb_a_vals[sheet]
        max_r = max(ws_b.max_row or 1, ws_a.max_row or 1)
        max_c = max(ws_b.max_column or 1, ws_a.max_column or 1)

        for cb, ca, cbv, cav in zip(
            _iter_cells(ws_b, max_r, max_c),
            _iter_cells(ws_a, max_r, max_c),
            _iter_cells(ws_bv, max_r, max_c),
            _iter_cells(ws_av, max_r, max_c),
        ):
            coord = _cell_coord(cb)
            b_formula = cb.value if isinstance(cb.value, str) and str(cb.value).startswith("=") else None
            a_formula = ca.value if isinstance(ca.value, str) and str(ca.value).startswith("=") else None
            if cb.data_type == "f":
                b_formula = "=" + str(cb.value) if cb.value and not str(cb.value).startswith("=") else str(cb.value or "")
            if ca.data_type == "f":
                a_formula = "=" + str(ca.value) if ca.value and not str(ca.value).startswith("=") else str(ca.value or "")

            b_val, a_val = cbv.value, cav.value
            val_eq = values_equal(b_val, a_val, opts, abs_tol, rel_tol)

            f_kind = classify_formula_change(
                b_formula if cb.data_type == "f" else (str(cb.value) if b_formula else None),
                a_formula if ca.data_type == "f" else (str(ca.value) if a_formula else None),
                b_val,
                a_val,
                val_eq,
            )

            sd = style_diff(style_snapshot(cb), style_snapshot(ca))
            kinds: list[str] = []
            if f_kind:
                kinds.append(f_kind)
            elif not val_eq:
                if b_val is None and a_val is not None and cb.value is None:
                    kinds.append("added")
                elif a_val is None and b_val is not None and ca.value is None:
                    kinds.append("removed")
                elif b_val is not None or a_val is not None or cb.value != ca.value:
                    kinds.append("value")
            if sd:
                kinds.append("style")

            if not kinds:
                continue

            kind = "+".join(dict.fromkeys(kinds)) if len(kinds) > 1 else kinds[0]
            change = CellChange(
                sheet=sheet,
                cell=coord,
                kind=kind,
                before={"value": b_val, "formula": b_formula, "style": sd.get("font") if sd else None},
                after={"value": a_val, "formula": a_formula, "style": sd.get("font") if sd else None},
            )
            if f_kind and (b_formula or a_formula):
                ast = formula_ast_diff(
                    str(b_formula) if b_formula else None,
                    str(a_formula) if a_formula else None,
                )
                change.before = {**(change.before or {}), "formulaAst": ast} if isinstance(change.before, dict) else change.before
            cells.append(change)
            bump(sheet, kind)

    struct_b = sheet_structure(wb_b)
    struct_a = sheet_structure(wb_a)
    structure = StructureDiff(sheets=diff_sheet_structure(struct_b, struct_a))
    named = NamedRangesDiff(**diff_named_ranges(wb_b, wb_a))
    hidden = HiddenDiff(**diff_hidden(wb_b, wb_a))
    tables = diff_tables(before_path, after_path)

    cf_changes = diff_conditional(
        extract_conditional_formatting(before_path),
        extract_conditional_formatting(after_path),
    )
    for c in cf_changes:
        cells.append(CellChange(**c))
        bump(c["sheet"], "style")

    charts_raw = diff_charts(
        extract_charts_from_xlsx(before_path),
        extract_charts_from_xlsx(after_path),
    )
    charts = [ChartDiff(sheet=c["sheet"], chart_id=c["chartId"], diff=c["diff"]) for c in charts_raw]

    wb_b.close()
    wb_a.close()
    wb_b_vals.close()
    wb_a_vals.close()

    return DiffResult(
        cells=cells,
        charts=charts,
        structure=structure,
        named_ranges=named,
        hidden=hidden,
        tables=tables,
        summary=DiffSummary(total_changes=len(cells), by_sheet=by_sheet),
    )
