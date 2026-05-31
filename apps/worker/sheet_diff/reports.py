from __future__ import annotations

import json
from html import escape
from pathlib import Path
from typing import Any

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from sheet_diff.models import CellChange, DiffResult

KIND_COLORS = {
    "value": "#FFE082",
    "formula": "#90CAF9",
    "style": "#CE93D8",
    "value+formula": "#FFAB91",
    "added": "#A5D6A7",
    "removed": "#EF9A9A",
}


def write_diff_workbook(result: DiffResult, before_path: Path, after_path: Path, out_path: Path) -> None:
    wb = load_workbook(after_path)
    fills = {
        k: PatternFill(start_color=c.replace("#", ""), end_color=c.replace("#", ""), fill_type="solid")
        for k, c in KIND_COLORS.items()
    }
    for change in result.cells:
        if change.sheet not in wb.sheetnames:
            continue
        ws = wb[change.sheet]
        try:
            cell = ws[change.cell]
            fill = fills.get(change.kind, fills["value"])
            cell.fill = fill
        except (ValueError, KeyError):
            pass
    wb.save(out_path)


def write_html_report(result: DiffResult, out_path: Path) -> None:
  sheets: dict[str, list[CellChange]] = {}
  for c in result.cells:
      sheets.setdefault(c.sheet, []).append(c)

  parts = [
      "<!DOCTYPE html><html><head><meta charset='utf-8'>",
      "<title>SheetDiff Report</title>",
      "<style>body{font-family:system-ui,sans-serif;margin:2rem;background:#fafafa}",
      "section{background:#fff;padding:1rem;margin:1rem 0;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08)}",
      "table{border-collapse:collapse;width:100%}th,td{border:1px solid #e0e0e0;padding:.5rem;text-align:left}",
      ".kind-value{background:#FFE082}.kind-formula{background:#90CAF9}.kind-style{background:#CE93D8}",
      ".kind-added{background:#A5D6A7}.kind-removed{background:#EF9A9A}</style></head><body>",
      "<h1>SheetDiff HTML Report</h1>",
      f"<p>Total changes: {result.summary.total_changes}</p>",
  ]
  for sheet, changes in sorted(sheets.items()):
      parts.append(f"<section><h2>{escape(sheet)} ({len(changes)} changes)</h2><table>")
      parts.append("<tr><th>Cell</th><th>Kind</th><th>Before</th><th>After</th></tr>")
      for ch in changes[:500]:
          parts.append(
              f"<tr class='kind-{escape(ch.kind)}'><td>{escape(ch.cell)}</td>"
              f"<td>{escape(ch.kind)}</td><td>{escape(str(ch.before)[:200])}</td>"
              f"<td>{escape(str(ch.after)[:200])}</td></tr>"
          )
      parts.append("</table></section>")
  if result.charts:
      parts.append("<section><h2>Chart changes</h2><pre>")
      parts.append(escape(json.dumps([c.model_dump() for c in result.charts], indent=2)[:8000]))
      parts.append("</pre></section>")
  parts.append("</body></html>")
  out_path.write_text("\n".join(parts), encoding="utf-8")


def write_json_report(result: DiffResult, out_path: Path) -> None:
    out_path.write_text(result.model_dump_json(indent=2), encoding="utf-8")


def write_summary_xlsx(result: DiffResult, out_path: Path) -> None:
    wb = xlsxwriter.Workbook(str(out_path))
    ws = wb.add_worksheet("Summary")
    ws.write_row(0, 0, ["Sheet", "Value", "Formula", "Style", "Total"])
    row = 1
    for sheet, counts in result.summary.by_sheet.items():
        ws.write_row(row, 0, [sheet, counts.get("value", 0), counts.get("formula", 0), counts.get("style", 0), counts.get("total", 0)])
        row += 1
    detail = wb.add_worksheet("Changes")
    detail.write_row(0, 0, ["Sheet", "Cell", "Kind", "Before", "After"])
    for i, c in enumerate(result.cells[:10000], start=1):
        detail.write_row(i, 0, [c.sheet, c.cell, c.kind, str(c.before)[:500], str(c.after)[:500]])
    wb.close()
