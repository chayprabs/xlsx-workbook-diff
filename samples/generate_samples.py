#!/usr/bin/env python3
"""Generate PRD sample workbook pairs for golden tests."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).parent


def invoice_pair():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Widget"
    ws["B2"] = 100.0
    ws["A3"] = "Service"
    ws["B3"] = 50.0
    wb.save(OUT / "invoice_before.xlsx")

    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Invoice"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Widget"
    ws["B2"] = 100.001
    ws["A3"] = "Service"
    ws["B3"] = 55.0
    ws["A4"] = "Tax"
    ws["B4"] = 15.5
    wb2.save(OUT / "invoice_after.xlsx")


def pricing_pair():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing"
    ws["A1"] = "SKU"
    ws["B1"] = "Price"
    for i, sku in enumerate(["A1", "B2", "C3"], start=2):
        ws[f"A{i}"] = sku
        ws[f"B{i}"] = round(10.0 * i * 1.0, 4)
    wb.save(OUT / "pricing_before.xlsx")

    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Pricing"
    ws["A1"] = "SKU"
    ws["B1"] = "Price"
    for i, sku in enumerate(["A1", "B2", "C3"], start=2):
        ws[f"A{i}"] = sku
        ws[f"B{i}"] = round(10.0 * i * 1.05, 4)
    summary = wb2.create_sheet("Summary")
    summary["A1"] = "Notes"
    summary["B1"] = "M2 pricing"
    wb2.save(OUT / "pricing_after.xlsx")


def financial_pair():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    for i, m in enumerate(["Jan", "Feb", "Mar"], start=2):
        ws[f"A{i}"] = m
        ws[f"B{i}"] = 1000 * i
    wb.defined_names.add(DefinedName("RevenueRange", attr_text="Report!$B$2:$B$4"))
    chart = BarChart()
    chart.title = "Revenue"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")
    wb.save(OUT / "financial_before.xlsx")

    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Report"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    for i, m in enumerate(["Jan", "Feb", "Mar"], start=2):
        ws[f"A{i}"] = m
        ws[f"B{i}"] = 1000 * i + (50 if i == 3 else 0)
    wb2.defined_names.add(DefinedName("RevenueRange", attr_text="Report!$B$2:$B$5"))
    chart = BarChart()
    chart.title = "Revenue Q1"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")
    wb2.save(OUT / "financial_after.xlsx")


def numerical_noise_pair():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1.0000000
    ws["A2"] = 2.0
    wb.save(OUT / "noise_before.xlsx")
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Data"
    ws["A1"] = 1.0000005
    ws["A2"] = 2.0000001
    wb2.save(OUT / "noise_after.xlsx")


if __name__ == "__main__":
    invoice_pair()
    pricing_pair()
    financial_pair()
    numerical_noise_pair()
    print("Samples generated in", OUT)
